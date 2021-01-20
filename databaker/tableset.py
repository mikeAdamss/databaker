# Devnote - this is a rewrite of the messytables excel table loader
# https://github.com/okfn/messytables/blob/master/messytables/excel.py
# to remove reliance on the outdated xlrd library

import sys
from datetime import datetime, time
import xlrd
from xlrd.biffh import XLRDError

from messytables.core import RowSet, TableSet, Cell, CoreProperties
from messytables.types import (StringType, IntegerType,
                               DateType, FloatType)
from messytables.error import ReadError
from messytables.compat23 import PY2

import openpyxl

class InvalidDateError(Exception):
    pass

def get_xls_type(openpyexcel_cell):
    """
    Map the raw type of the openpyexcel cell to a messytable cell type
    """
    if isinstance(openpyexcel_cell, float):
        return FloatType()
    if isinstance(openpyexcel_cell, time):
        return DateType(None)
    if isinstance(openpyexcel_cell, int):
        return IntegerType()
    return StringType()

class GSSExcelTableSet(TableSet):
    """An excel workbook wrapper object.
    """

    def __init__(self, fileobj=None, filename=None, window=None,
                 encoding=None, with_formatting_info=True, **kw):

        def get_workbook():
            if fileobj is not None:
                return openpyxl.load_workbook(fileobj)
            return openpyxl.load_workbook(filename)

        self.window = window

        if not filename and not fileobj:
            raise Exception('You must provide one of filename or fileobj')

        self.workbook = get_workbook()


    def make_tables(self):
        """ Return the sheets in the workbook. """
        return [GSSTableRowSet(name, self.workbook[name], self.window)
                for name in self.workbook.sheetnames]


class GSSTableRowSet(RowSet):
    """ Excel support for a single sheet in the excel workbook. Unlike
    the CSV row set this is not a streaming operation. """

    def __init__(self, name, sheet, window=None):
        self.name = name
        self.sheet = sheet
        self.window = window or 1000
        super(GSSTableRowSet, self).__init__(typed=True)

    def raw(self, sample=False):
        for row_no, row in enumerate(self.sheet.iter_rows()):
            row_of_cells = []
            for col_no, openpyexcel_cell in enumerate(row):
                row_of_cells.append(
                    XLSCell.from_openpyexcel(openpyexcel_cell,
                            self.sheet, col_no, row_no))
            yield row_of_cells


class XLSCell(Cell):

    @staticmethod
    def from_openpyexcel(openpyexcel_cell, sheet, col, row):
        value = openpyexcel_cell.value
        cell_type = get_xls_type(openpyexcel_cell)

        # time parsing
        if cell_type == DateType(None):
            # TODO - possible? not sure we care
            pass

        messy_cell = XLSCell(value, type=cell_type)
        messy_cell.sheet = sheet
        messy_cell.xlrd_cell = openpyexcel_cell  # hmmmmmmmmm
        messy_cell.xlrd_pos = (row, col)
        return messy_cell

    # @property
    # def topleft(self):
    #     return self.properties.topleft

    @property
    def properties(self):
        return GSSXLSProperties(self)

class GSSXLSProperties(CoreProperties):
    #KEYS = ['bold', 'size', 'italic', 'font_name', 'strikeout', 'underline',
    #        'font_colour', 'background_colour', 'any_border', 'all_border',
    #        'richtext', 'blank', 'a_date', 'formatting_string']
    
    def __init__(self, cell):
        self.cell = cell
        self.merged = {}

    def get_bold(self):
        return self.cell.font.bold  # the openpyexcel way of expressing bold

    # @property
    #def xf(self):
    #    return self.cell.sheet.book.xf_list[self.cell.xlrd_cell.xf_index]

    # @property
    # def font(self):
    #    return self.cell.sheet.book.font_list[self.xf.font_index]

    # @property
    # def formatting(self):
    #    return self.cell.sheet.book.format_map[self.xf.format_key]

    # @property
    # def rich(self):
    #    """returns a tuple of character position, font number which starts at that position
    #    https://secure.simplistix.co.uk/svn/xlrd/trunk/xlrd/doc/xlrd.html?p=4966#sheet.Sheet.rich_text_runlist_map-attribute"""
    #    return self.cell.sheet.rich_text_runlist_map.get(self.cell.xlrd_pos, None)

    # def raw_span(self, always=False):
    #    """return the bounding box of the cells it's part of.
    #     https://secure.simplistix.co.uk/svn/xlrd/trunk/xlrd/doc/xlrd.html?p=4966#sheet.Sheet.merged_cells-attribute"""
    #    row, col = self.cell.xlrd_pos
    #    for box in self.cell.sheet.merged_cells:
    #        rlo, rhi, clo, chi = box
    #        # note the high indexes are NOT inclusive!
    #        rhi = rhi - 1
    #        chi = chi - 1
    #        if row >= rlo and row <= rhi and col >= clo and col <= chi:
    #            return rlo, rhi, clo, chi
    #    if always:
    #        return (row, row, col, col)
    #    else:
    #        return None

    # @property
    # def topleft(self):
    #    span = self.raw_span()
    #    if span is None:
    #        return True  # is a single cell
    #    else:
    #        rlo, _, clo, _ = span
    #        return (rlo, clo) == self.cell.xlrd_pos

    # def get_formatting_string(self):
    #    return self.formatting.format_str

    # def get_a_date(self):
    #    return self.formatting.type == 1

    # def get_richtext(self):  # TODO - get_rich_fragments
    #    return bool(self.rich)

    # def get_bold(self):
    #    return self.font.weight > 500

    # def get_size(self):
    #    """in pixels"""
    #    return self.font.height / 20.0

    # def get_italic(self):
    #    return bool(self.font.italic)

    # def get_font_name(self):
    #    return self.font.name

    # def get_strikeout(self):
    #    return bool(self.font.struck_out)

    # def get_underline(self):
    #    return self.font.underline_type > 0

    # def get_font_colour(self):
        # TODO
    #    return self.font.color_index ## more lookup required

    # def get_blank(self):
    #    """Note that cells might not exist at all.
    #       Behaviour for spanned cells might be complicated: hence this function"""
    #    return self.cell.value == ''

    # def get_background_colour(self):
    #    return self.xf.background.background_color_index ## more lookup required

    # def get_any_border(self):
    #    b = self.xf.border
    #    return b.top_line_style > 0 or b.bottom_line_style > 0 or \
    #           b.left_line_style > 0 or b.right_line_style > 0

    # def get_all_border(self):
    #     b = self.xf.border
    #     return b.top_line_style > 0 and b.bottom_line_style > 0 and \
    #           b.left_line_style > 0 and b.right_line_style > 0
